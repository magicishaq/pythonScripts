import os
from openpyxl import load_workbook
os.chdir('C:\\Users\\ikhan\\Documents\\learningPython\\excel') #change location to where the spreadsheet is

workbook = load_workbook(filename='Prices.xlsx')
workbook.sheetnames
sheet = workbook.active

price = sheet['A2'].value #dogfood

item = sheet.cell(row=1, column=1).value #another way to grab a cell

