from openpyxl import Workbook
import os

workbook = Workbook()
sheet = workbook.active #gets the first sheet

checker = sheet['A1'].value == None

sheet['A1'] = 'Inventory'
sheet['A2'] = 'Vampires spray'
sheet['B1'] = 'Quantity'
sheet['B2'] = 200

#currently this spreadsheet only exists on the file
#creating a new sheet
secondSheet = workbook.create_sheet('Holidays')
#changing name of the sheet
sheet.title = 'InventorySheet'
#creating a sheet at an index
workbook.create_sheet(index=0, title="firstSheet")
#saving the spreadsheet
os.chdir('C:\\Users\\ikhan\\Documents\\learningPython\\excel') #where it will be saved
workbook.save('Inventory.xlsx') #filename




print(workbook.get_sheet_names())


